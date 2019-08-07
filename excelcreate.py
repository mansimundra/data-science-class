from openpyxl import Workbook,load_workbook
wb=Workbook()
filepath="C:\\Users\\Rajesh\\Documents\\createfile.xlsx"
# wb.save(filepath)

# # import load_workbook
# from openpyxl import load_workbook
# # set file path
# filepath="C:\\Users\\Rajesh\\Documents\\createfile.xlsx"
# # load demo.xlsx
# wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# set value for cell A1=1
sheet['A1'] = 'NAME'
sheet['B1'] = 'AGE'
sheet['C1'] = 'FIELD'
# sheet.cell(row=2, column=2).value = 8
data=[['abc',29,'indore'],['BCD',22,'TEXAS']]
for row in data:
    sheet.append(row)
print(sheet)
wb.save(filepath)
# wb.save(filepath)
# for i in range(1,len(data)):
#     for j in range (1,len(data[i])):
#         sheet.cell(row=(i+1), column=i).value=data[i][j]
