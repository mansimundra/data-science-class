from openpyxl import Workbook,load_workbook
filename="C:\\Users\\Rajesh\\PycharmProjects\\newpro\\new.xlsx"
wb=load_workbook(filename)
sheetnames=wb.sheetnames
print(sheetnames)
mysheet1=wb.create_sheet("My sheet 1",0)
print(mysheet1)
print(wb.sheetnames)
sheet=wb[wb.sheetnames[0]]
sheet['A1']='NAME'
sheet['B1']='AGE'
sheet['C1']='FIELD'
wb.save(filename)
data=[('MANSI',22,'IT'),('IRA',23,'FINANCE')]
for rows in data:
    sheet.append(rows)
